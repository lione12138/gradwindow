from __future__ import annotations

import re
from collections.abc import Callable
from io import BytesIO

import httpx
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from ..http_client import DEFAULT_USER_AGENT
from .base import DiscoveredCatalog, DiscoveredProgramme, DiscoveredWindow, Fetcher
from .official_catalog import normalise, slug

GUIDE_URL = "https://intl.csu.edu.cn/info/1140/3776.htm"
CATALOG_URL = (
    "https://intl.csu.edu.cn/system/_content/download.jsp?"
    "urltype=news.DownloadAttachUrl&owner=1667182464&wbfileid=16620166"
)
APPLICATION_URL = "https://csu.17gz.org/"

CatalogueEntries = tuple[tuple[str, str, str, str], ...]
CatalogueFetcher = Callable[[str], CatalogueEntries]


class CentralSouthAdapter:
    university_id = "central-south-university"
    catalog_url = CATALOG_URL
    application_url = APPLICATION_URL
    intake = "Autumn 2026"
    application_opens_at_basis = "missing"
    replace_pending_candidates = True
    window_watch_urls = (GUIDE_URL,)
    known_programme_window_scope_type = "programme-group"
    catalogue_limitation_reason = (
        "Central South's official 2026 guide says applications run 'from now' "
        "to exact closing dates. Because 'from now' is not an exact opening "
        "date, those dates remain review guidance rather than publishable windows."
    )

    def __init__(
        self,
        minimum_expected_programmes: int = 150,
        maximum_expected_programmes: int = 175,
        catalogue_fetcher: CatalogueFetcher | None = None,
    ) -> None:
        self.minimum_expected_programmes = minimum_expected_programmes
        self.maximum_expected_programmes = maximum_expected_programmes
        self.catalogue_fetcher = catalogue_fetcher or _fetch_catalogue

    def parse_catalog_from_fetcher(self, fetcher: Fetcher) -> DiscoveredCatalog:
        programmes = _programmes(self.catalogue_fetcher(CATALOG_URL))
        if not (
            self.minimum_expected_programmes
            <= len(programmes)
            <= self.maximum_expected_programmes
        ):
            raise ValueError(
                f"Central South catalogue contained {len(programmes)} master's "
                f"programmes; expected {self.minimum_expected_programmes}-"
                f"{self.maximum_expected_programmes}"
            )
        closings = _closing_dates(fetcher(GUIDE_URL))
        programmes.extend(_review_groups(closings))
        return DiscoveredCatalog(application_opens_at=None, programmes=programmes)


def _programmes(entries: CatalogueEntries) -> list[DiscoveredProgramme]:
    programmes: dict[tuple[str, str, str], DiscoveredProgramme] = {}
    for raw_faculty, raw_name, _research_field, raw_medium in entries:
        faculty = normalise(raw_faculty)
        name = normalise(raw_name)
        medium = normalise(raw_medium).replace("/", " and ")
        if not faculty or not name or not medium:
            continue
        key = (faculty.casefold(), name.casefold(), medium.casefold())
        programmes[key] = DiscoveredProgramme(
            id=f"central-south-{slug(faculty)}-{slug(name)}-{slug(medium)}",
            name=f"{name} ({medium}-medium)",
            degree_type="Master",
            faculty=faculty,
            department=faculty,
            source_url=CATALOG_URL,
            application_url=APPLICATION_URL,
            windows=[],
            deadline_text=(
                "Programme is listed in Central South University's official "
                "2026 international master's workbook. Research fields are "
                "deduplicated at major and teaching-language scope."
            ),
            parse_status="no-deadline",
            retrieval_method="official-2026-international-master-catalogue-xlsx",
            evidence_quality="official-full-text",
        )
    return sorted(
        programmes.values(),
        key=lambda item: (item.faculty.casefold(), item.name.casefold()),
    )


def _review_groups(closings: dict[str, str]) -> list[DiscoveredProgramme]:
    definitions = (
        (
            "central-south-cgs-high-level-graduate-admissions",
            "Chinese Government Scholarship high-level graduate admissions",
            "Chinese Government Scholarship high-level graduate round",
            "chinese-government-scholarship",
            closings["scholarship"],
        ),
        (
            "central-south-scholarship-self-sponsored-admissions",
            "University scholarship and self-sponsored admissions",
            "University scholarship and self-sponsored round",
            "international-students",
            closings["general"],
        ),
    )
    return [
        DiscoveredProgramme(
            id=programme_id,
            name=name,
            degree_type="Master/Doctoral",
            faculty="School of International Education",
            department="School of International Education",
            source_url=GUIDE_URL,
            application_url=APPLICATION_URL,
            windows=[
                DiscoveredWindow(
                    round=round_name,
                    applicant_categories=[category],
                    opens_at=None,
                    closes_at=closes_at,
                    intake="Autumn 2026",
                    source_url=GUIDE_URL,
                    opens_at_basis="missing",
                )
            ],
            deadline_text=(
                "The official 2026 guide gives this exact closing date but "
                "describes the opening only as 'from now', so it remains "
                "review guidance and cannot be published as an exact window."
            ),
            parse_status="incomplete",
            retrieval_method="official-2026-international-graduate-guide-html",
            evidence_quality="official-full-text",
        )
        for programme_id, name, round_name, category, closes_at in definitions
    ]


def _closing_dates(html: str) -> dict[str, str]:
    text = normalise(BeautifulSoup(html, "html.parser").get_text(" ", strip=True))
    compact = re.sub(r"(?<=\d)\s+(?=\d)", "", text)
    if not re.search(r"From now on to February 15\s*,?\s*2026", compact, re.I):
        raise ValueError("Central South guide lacked its scholarship closing date")
    if not re.search(r"From now on to May 31\s*,?\s*2026", compact, re.I):
        raise ValueError("Central South guide lacked its general closing date")
    return {"scholarship": "2026-02-15", "general": "2026-05-31"}


def _fetch_catalogue(url: str) -> CatalogueEntries:
    headers = {
        "User-Agent": DEFAULT_USER_AGENT,
        "Accept": (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
            "application/octet-stream,*/*;q=0.8"
        ),
    }
    with httpx.Client(follow_redirects=True, timeout=90, headers=headers) as client:
        guide_response = client.get(GUIDE_URL)
        guide_response.raise_for_status()
        response = client.get(url, headers={"Referer": GUIDE_URL})
        response.raise_for_status()
        content = response.content
    if len(content) > 1_000_000 or not content.startswith(b"PK"):
        raise ValueError("Central South catalogue did not return a bounded workbook")
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        if not workbook.worksheets or "Master" not in str(
            workbook.worksheets[0]["A1"].value
        ):
            raise ValueError("Central South workbook lacked its master's worksheet")
        entries = []
        for row in workbook.worksheets[0].iter_rows(min_row=3, values_only=True):
            faculty, name, research_field, medium = row[1], row[3], row[5], row[7]
            if faculty and name and medium:
                entries.append(
                    (
                        normalise(faculty),
                        normalise(name),
                        normalise(research_field or ""),
                        normalise(medium),
                    )
                )
        return tuple(entries)
    finally:
        workbook.close()
