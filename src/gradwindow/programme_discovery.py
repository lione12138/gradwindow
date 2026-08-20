from __future__ import annotations

import hashlib
import json
import re
from collections.abc import Callable
from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from .content import deadline_signal_text
from .http_client import DEFAULT_USER_AGENT, fetch_page
from .intakes import parse_intake_details
from .io import read_json, write_json
from .monitor import extract_fetched_text
from .paths import (
    APPLICANT_CATEGORIES_PATH,
    APPLICATIONS_PATH,
    PROGRAMME_CANDIDATES_PATH,
    PROGRAMME_CATALOG_STATE_PATH,
    PROGRAMS_PATH,
    WINDOW_CANDIDATES_PATH,
)
from .predictions import official_cycle_key
from .programme_adapters.base import ProgrammeAdapter
from .programme_windows import (
    has_official_exact_window,
    known_programme_window_candidates,
)

WATCHED_SOURCE_FINGERPRINT_VERSION = 2


def fetch_catalog(url: str) -> str:
    page = fetch_page(
        url,
        user_agent=DEFAULT_USER_AGENT,
        timeout=30,
        max_bytes=8_000_000,
        accept=(
            "text/html,application/xhtml+xml,application/xml;q=0.9,"
            "text/xml;q=0.8,*/*;q=0.7"
        ),
    )
    if (
        "spreadsheetml.sheet" in page.content_type.lower()
        or page.final_url.lower().split("?", 1)[0].endswith(".xlsx")
    ):
        return _xlsx_payload(page.raw_bytes)
    return extract_fetched_text(page)


def _xlsx_payload(raw_bytes: bytes) -> str:
    workbook = load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)
    worksheets = []
    for sheet in workbook.worksheets:
        rows = [
            [_json_cell(value) for value in row]
            for row in sheet.iter_rows(values_only=True)
        ]
        worksheets.append({"name": sheet.title, "rows": rows})
    workbook.close()
    return json.dumps({"worksheets": worksheets}, ensure_ascii=False)


def _json_cell(value):
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def discover_programmes(
    adapter: ProgrammeAdapter,
    *,
    programs_path: Path = PROGRAMS_PATH,
    applications_path: Path = APPLICATIONS_PATH,
    candidates_path: Path = PROGRAMME_CANDIDATES_PATH,
    window_candidates_path: Path | None = None,
    state_path: Path = PROGRAMME_CATALOG_STATE_PATH,
    applicant_categories_path: Path = APPLICANT_CATEGORIES_PATH,
    fetcher: Callable[[str], str] = fetch_catalog,
    dry_run: bool = False,
) -> dict:
    checked_at = datetime.now(timezone.utc).isoformat()
    watched_urls = set(getattr(adapter, "window_watch_urls", ()))
    watched_source_hashes: dict[str, str] = {}

    def tracking_fetcher(url: str) -> str:
        content = fetcher(url)
        if url in watched_urls:
            watched_source_hashes[url] = _source_hash(content)
        return content

    catalog = adapter.parse_catalog_from_fetcher(tracking_fetcher)
    _validate_recurring_applicant_categories(catalog, applicant_categories_path)
    programs_payload = read_json(programs_path)
    known_programmes = {
        item["id"]: item
        for item in programs_payload.get("programs", [])
        if item.get("universityId") == adapter.university_id
    }
    known_ids = set(known_programmes)
    candidates_payload = read_json(
        candidates_path,
        {
            "meta": {
                "description": (
                    "Automatically discovered programme candidates awaiting "
                    "manual review. This file is not published."
                )
            },
            "items": [],
        },
    )
    existing = {item["id"]: item for item in candidates_payload.get("items", [])}
    if adapter.replace_pending_candidates:
        existing = {
            candidate_id: item
            for candidate_id, item in existing.items()
            if item.get("universityId") != adapter.university_id
            or (
                item.get("status", "pending") != "pending"
                and item.get("type") != "known-programme-recurring-policy"
            )
        }
    if window_candidates_path is None:
        window_candidates_path = (
            WINDOW_CANDIDATES_PATH
            if candidates_path == PROGRAMME_CANDIDATES_PATH
            else candidates_path.with_name("window-candidates.json")
        )
    applications = read_json(applications_path, {"applications": []}).get(
        "applications", []
    )
    window_candidates_payload = read_json(
        window_candidates_path,
        {
            "meta": {
                "description": (
                    "Internal exact-window candidates awaiting manual review. "
                    "This file is never published to the static site."
                )
            },
            "items": [],
        },
    )
    existing_window_candidates = {
        item["id"]: item for item in window_candidates_payload.get("items", [])
    }
    original_window_candidate_items = window_candidates_payload.get("items", [])
    if adapter.replace_pending_candidates:
        existing_window_candidates = {
            candidate_id: item
            for candidate_id, item in existing_window_candidates.items()
            if item.get("universityId") != adapter.university_id
            or item.get("status", "pending") != "pending"
        }
    original_window_candidates_by_id = {
        item["id"]: item for item in window_candidates_payload.get("items", [])
    }
    applications_by_cycle = {
        official_cycle_key(item): item
        for item in applications
        if item.get("universityId") == adapter.university_id
    }
    application_ids = {item["id"] for item in applications}
    created = 0
    created_guidance_candidates = 0
    created_window_candidates = 0
    changed_window_candidates = 0
    for programme in catalog.programmes:
        if programme.id in known_ids:
            for candidate in known_programme_window_candidates(
                adapter,
                programme,
                known_programmes[programme.id],
                catalog.application_opens_at,
                applications_by_cycle,
                application_ids,
                checked_at,
            ):
                previous = existing_window_candidates.get(candidate["id"])
                if previous is not None:
                    candidate["status"] = previous.get("status", "pending")
                    candidate["detectedAt"] = previous.get("detectedAt", checked_at)
                    candidate["record"]["verifiedAt"] = previous.get("record", {}).get(
                        "verifiedAt", candidate["record"]["verifiedAt"]
                    )
                    for key in ("reviewedAt", "reviewedBy", "reviewNotes"):
                        if key in previous:
                            candidate[key] = previous[key]
                else:
                    created_window_candidates += 1
                    if candidate["type"] == "adapter-window-change":
                        changed_window_candidates += 1
                existing_window_candidates[candidate["id"]] = candidate
            guidance = _known_programme_guidance_candidate(
                adapter,
                programme,
                catalog.application_opens_at,
                checked_at,
            )
            if guidance is not None:
                previous = existing.get(guidance["id"])
                if previous is None:
                    if guidance["type"] != "known-programme-recurring-policy":
                        created_guidance_candidates += 1
                else:
                    guidance["status"] = previous.get("status", "pending")
                    guidance["detectedAt"] = previous.get("detectedAt", checked_at)
                    for key in ("reviewedAt", "reviewedBy", "reviewNotes"):
                        if key in previous:
                            guidance[key] = previous[key]
                existing[guidance["id"]] = guidance
            continue
        candidate_id = f"new-programme:{programme.id}"
        previous = existing.get(candidate_id)
        candidate = _candidate_record(
            adapter,
            programme,
            catalog.application_opens_at,
            checked_at,
        )
        if previous is None:
            created += 1
        else:
            candidate["status"] = previous.get("status", "pending")
            candidate["detectedAt"] = previous.get("detectedAt", checked_at)
            for key in ("reviewedAt", "reviewedBy", "reviewNotes"):
                if key in previous:
                    candidate[key] = previous[key]
        existing[candidate_id] = candidate

    original_candidate_items = candidates_payload.get("items", [])
    unrelated_items = [
        existing[item["id"]]
        for item in original_candidate_items
        if item.get("universityId") != adapter.university_id and item["id"] in existing
    ]
    target_items = sorted(
        (
            item
            for item in existing.values()
            if item.get("universityId") == adapter.university_id
        ),
        key=lambda item: (item.get("status") != "pending", item["id"]),
    )
    items = unrelated_items + target_items
    snapshot_items = {
        programme.id: {
            "name": programme.name,
            "degreeType": programme.degree_type,
            "faculty": programme.faculty,
            "department": programme.department,
            "parseStatus": programme.parse_status,
            "deadlineHash": _hash(
                json.dumps(
                    [
                        {
                            "round": window.round,
                            "opensAt": window.opens_at,
                            "closesAt": window.closes_at,
                            "applicantCategories": window.applicant_categories,
                            "sourceUrl": window.source_url,
                            "opensAtBasis": window.opens_at_basis,
                        }
                        for window in programme.windows
                    ],
                    sort_keys=True,
                )
            ),
        }
        for programme in catalog.programmes
    }
    window_snapshot_items: dict[str, dict] = {}
    for programme in catalog.programmes:
        identities: dict[str, int] = {}
        for window in sorted(
            programme.windows,
            key=lambda item: (
                item.intake or adapter.intake,
                item.round,
                ",".join(sorted(item.applicant_categories)),
                item.opens_at or "",
                item.closes_at,
            ),
        ):
            base_identity = "::".join(
                (
                    programme.id,
                    window.intake or adapter.intake,
                    window.round,
                    ",".join(sorted(window.applicant_categories)) or "all",
                )
            )
            identities[base_identity] = identities.get(base_identity, 0) + 1
            identity = base_identity
            if identities[base_identity] > 1:
                identity = f"{base_identity}::{identities[base_identity]}"
            window_snapshot_items[identity] = {
                "programmeId": programme.id,
                "intake": window.intake or adapter.intake,
                "opensAt": window.opens_at,
                "closesAt": window.closes_at,
                "sourceUrl": window.source_url or programme.source_url,
                "opensAtBasis": window.opens_at_basis,
            }
    state_payload = read_json(state_path, {"meta": {}, "universities": {}})
    previous_state = state_payload.get("universities", {}).get(
        adapter.university_id, {}
    )
    previous_programmes = previous_state.get("programmes")
    previous_windows = previous_state.get("windows")
    has_record_baseline = isinstance(previous_programmes, dict) and isinstance(
        previous_windows, dict
    )
    if has_record_baseline:
        previous_programme_ids = set(previous_programmes)
        current_programme_ids = set(snapshot_items)
        previous_window_ids = set(previous_windows)
        current_window_ids = set(window_snapshot_items)
        record_diff = {
            "previous": {
                "programmes": len(previous_programmes),
                "windows": len(previous_windows),
            },
            "current": {
                "programmes": len(snapshot_items),
                "windows": len(window_snapshot_items),
            },
            "disappearedProgrammeIds": sorted(
                previous_programme_ids - current_programme_ids
            ),
            "addedProgrammeIds": sorted(current_programme_ids - previous_programme_ids),
            "changedProgrammeIds": sorted(
                programme_id
                for programme_id in previous_programme_ids & current_programme_ids
                if previous_programmes[programme_id] != snapshot_items[programme_id]
            ),
            "disappearedWindowIds": sorted(previous_window_ids - current_window_ids),
            "addedWindowIds": sorted(current_window_ids - previous_window_ids),
            "changedWindowIds": sorted(
                window_id
                for window_id in previous_window_ids & current_window_ids
                if previous_windows[window_id] != window_snapshot_items[window_id]
            ),
        }
        disappeared_window_details = {
            window_id: previous_windows[window_id]
            for window_id in record_diff["disappearedWindowIds"]
        }
    else:
        record_diff = {
            "previous": {
                "programmes": int(previous_state.get("itemCount", 0)),
                "windows": int(previous_state.get("observedWindowCount", 0)),
            },
            "current": {
                "programmes": len(snapshot_items),
                "windows": len(window_snapshot_items),
            },
            "disappearedProgrammeIds": [],
            "addedProgrammeIds": [],
            "changedProgrammeIds": [],
            "disappearedWindowIds": [],
            "addedWindowIds": [],
            "changedWindowIds": [],
        }
        disappeared_window_details = {}
    observed_window_count = sum(
        len(programme.windows) for programme in catalog.programmes
    )
    exact_window_count = sum(
        _is_official_exact_window(adapter, catalog, window)
        for programme in catalog.programmes
        for window in programme.windows
    )
    recurring_policy_window_count = sum(
        _is_official_recurring_policy_window(adapter, catalog, window)
        for programme in catalog.programmes
        for window in programme.windows
    )
    window_counts_by_cycle = _window_counts_by_cycle(adapter, catalog)
    missing_opening_date_count = sum(
        not _has_official_exact_opening(adapter, catalog, window)
        and not _is_official_recurring_policy_window(adapter, catalog, window)
        for programme in catalog.programmes
        for window in programme.windows
    )
    window_fingerprint = _hash(
        json.dumps(
            {
                programme.id: [
                    {
                        "round": window.round,
                        "opensAt": window.opens_at,
                        "closesAt": window.closes_at,
                        "intake": window.intake,
                        "applicantCategories": window.applicant_categories,
                        "sourceUrl": window.source_url,
                        "opensAtBasis": window.opens_at_basis,
                    }
                    for window in programme.windows
                ]
                for programme in catalog.programmes
            },
            sort_keys=True,
        )
    )
    watched_source_fingerprint = (
        _hash(json.dumps(watched_source_hashes, sort_keys=True))
        if watched_source_hashes
        else None
    )
    watched_source_fingerprint_version = (
        WATCHED_SOURCE_FINGERPRINT_VERSION if watched_source_hashes else None
    )
    programmes_without_deadlines = sum(
        not programme.windows for programme in catalog.programmes
    )
    programmes_needing_review = sum(
        programme.parse_status not in {"parsed", "recurring-policy"}
        for programme in catalog.programmes
    )
    window_status = _window_status(
        exact_window_count,
        recurring_policy_window_count,
        observed_window_count,
        missing_opening_date_count,
        programmes_without_deadlines,
    )
    limitation_reason = _limitation_reason(
        exact_window_count,
        recurring_policy_window_count,
        observed_window_count,
        missing_opening_date_count,
        programmes_without_deadlines,
    )
    catalogue_status = getattr(adapter, "catalogue_status", "ok")
    catalogue_limitation_reason = getattr(adapter, "catalogue_limitation_reason", None)
    if catalogue_limitation_reason:
        limitation_reason = " ".join(
            reason
            for reason in (catalogue_limitation_reason, limitation_reason)
            if reason
        )
    state_payload.setdefault("universities", {})[adapter.university_id] = {
        "sourceUrl": adapter.catalog_url,
        "checkedAt": checked_at,
        "itemCount": len(catalog.programmes),
        "applicationOpensAt": catalog.application_opens_at,
        "catalogHash": _hash(json.dumps(snapshot_items, sort_keys=True)),
        "catalogueStatus": catalogue_status,
        "windowStatus": window_status,
        "observedWindowCount": observed_window_count,
        "exactWindowCount": exact_window_count,
        "recurringPolicyWindowCount": recurring_policy_window_count,
        "windowCountsByCycle": window_counts_by_cycle,
        "missingOpeningDateCount": missing_opening_date_count,
        "programmesWithoutDeadlines": programmes_without_deadlines,
        "programmesNeedingReview": programmes_needing_review,
        "limitationReason": limitation_reason,
        "lastSuccessfulAt": checked_at,
        "windowFingerprint": window_fingerprint,
        "watchedWindowSourceHash": watched_source_fingerprint,
        "watchedWindowSourceFingerprintVersion": (watched_source_fingerprint_version),
        "programmes": snapshot_items,
        "windows": window_snapshot_items,
        "adapterWarnings": catalog.warnings,
    }
    state_payload["meta"] = {
        "updatedAt": checked_at,
        "description": "Latest official programme-catalog discovery snapshots.",
    }

    if not dry_run:
        candidates_payload["items"] = items
        candidates_payload.setdefault("meta", {})["updatedAt"] = checked_at
        write_json(candidates_path, candidates_payload)
        unrelated_window_candidate_items = [
            existing_window_candidates[item["id"]]
            for item in original_window_candidate_items
            if item.get("universityId") != adapter.university_id
            and item["id"] in existing_window_candidates
        ]
        target_window_candidate_items = sorted(
            (
                item
                for item in existing_window_candidates.values()
                if item.get("universityId") == adapter.university_id
            ),
            key=lambda item: (item.get("status") != "pending", item["id"]),
        )
        window_candidate_items = (
            unrelated_window_candidate_items + target_window_candidate_items
        )
        if existing_window_candidates != original_window_candidates_by_id:
            window_candidates_payload["items"] = window_candidate_items
            window_candidates_payload.setdefault("meta", {})["updatedAt"] = checked_at
            write_json(window_candidates_path, window_candidates_payload)
        write_json(state_path, state_payload)

    return {
        "status": "ok",
        "universityId": adapter.university_id,
        "sourceUrl": adapter.catalog_url,
        "checkedAt": checked_at,
        "catalogProgrammes": len(catalog.programmes),
        "previousCatalogProgrammes": previous_state.get("itemCount"),
        "knownProgrammes": len(known_ids),
        "newCandidates": created,
        "newGuidanceCandidates": created_guidance_candidates,
        "publishedRecurringPolicyRecords": sum(
            item.get("type") == "known-programme-recurring-policy"
            and item.get("status") == "published"
            and item.get("universityId") == adapter.university_id
            for item in items
        ),
        "newWindowCandidates": created_window_candidates,
        "changedWindowCandidates": changed_window_candidates,
        "pendingWindowCandidates": sum(
            item.get("status", "pending") == "pending"
            and item.get("universityId") == adapter.university_id
            for item in existing_window_candidates.values()
        ),
        "pendingCandidates": sum(
            item.get("status", "pending") == "pending"
            and item.get("universityId") == adapter.university_id
            and item.get("type") != "known-programme-recurring-policy"
            for item in items
        ),
        "pendingGuidanceCandidates": sum(
            item.get("status", "pending") == "pending"
            and item.get("universityId") == adapter.university_id
            and item.get("type") == "known-programme-window-guidance"
            for item in items
        ),
        "catalogueStatus": catalogue_status,
        "windowStatus": window_status,
        "observedWindowCount": observed_window_count,
        "exactWindowCount": exact_window_count,
        "recurringPolicyWindowCount": recurring_policy_window_count,
        "windowCountsByCycle": window_counts_by_cycle,
        "missingOpeningDateCount": missing_opening_date_count,
        "programmesWithoutDeadlines": programmes_without_deadlines,
        "programmesNeedingReview": programmes_needing_review,
        "limitationReason": limitation_reason,
        "adapterWarnings": catalog.warnings,
        "recordDiff": record_diff,
        "windowRemovalAssessmentAvailable": has_record_baseline,
        "disappearedWindowDetails": disappeared_window_details,
        "windowFingerprint": window_fingerprint,
        "watchedWindowSourceHash": watched_source_fingerprint,
        "watchedWindowSourceFingerprintVersion": (watched_source_fingerprint_version),
        "dryRun": dry_run,
    }


def _window_counts_by_cycle(adapter, catalog) -> dict[str, dict]:
    counts: dict[str, dict] = {}
    for programme in catalog.programmes:
        for window in programme.windows:
            intake = window.intake or adapter.intake
            cycle_key = _intake_cycle_key(intake)
            entry = counts.setdefault(
                cycle_key,
                {
                    "intakes": set(),
                    "observedWindowCount": 0,
                    "exactWindowCount": 0,
                    "recurringPolicyWindowCount": 0,
                },
            )
            entry["intakes"].add(intake)
            entry["observedWindowCount"] += 1
            entry["exactWindowCount"] += int(
                _is_official_exact_window(adapter, catalog, window)
            )
            entry["recurringPolicyWindowCount"] += int(
                _is_official_recurring_policy_window(adapter, catalog, window)
            )
    return {
        cycle_key: {**entry, "intakes": sorted(entry["intakes"])}
        for cycle_key, entry in sorted(counts.items())
    }


def _intake_cycle_key(intake: str) -> str:
    try:
        details = parse_intake_details(intake)
    except ValueError:
        normalised = re.sub(r"[^a-z0-9]+", "-", intake.lower()).strip("-")
        return f"label:{normalised or 'unknown'}"
    academic_year = (
        f"{details['cycleYear']}-{details['academicYearEnd']}"
        if details.get("academicYearEnd")
        else str(details["cycleYear"])
    )
    month = details.get("startMonth")
    return f"{academic_year}:{details['term']}:{month or 0:02d}"


def _validate_recurring_applicant_categories(
    catalog,
    applicant_categories_path: Path,
) -> None:
    allowed = {
        item["id"]
        for item in read_json(applicant_categories_path).get("categories", [])
        if item.get("id")
    }
    for programme in catalog.programmes:
        for window in programme.windows:
            if window.opens_at_basis != "official-recurring-policy":
                continue
            unknown = sorted(set(window.applicant_categories) - allowed)
            if unknown:
                raise ValueError(
                    f"{programme.id}: unknown recurring-policy applicant "
                    f"categories: {', '.join(unknown)}"
                )


def _known_programme_guidance_candidate(
    adapter,
    programme,
    shared_opens_at: str | None,
    detected_at: str,
) -> dict | None:
    candidate = _candidate_record(
        adapter,
        programme,
        shared_opens_at,
        detected_at,
    )
    if not candidate["windows"]:
        return None
    unresolved_windows = [
        window
        for window in candidate["windows"]
        if not has_official_exact_window(window)
    ]
    if candidate["windows"] and not unresolved_windows:
        return None
    candidate["windows"] = unresolved_windows
    if unresolved_windows and all(
        window.get("opensAtBasis") == "official-recurring-policy"
        for window in unresolved_windows
    ):
        candidate["id"] = f"known-programme-recurring:{programme.id}"
        candidate["type"] = "known-programme-recurring-policy"
        candidate["status"] = "published"
        candidate["reviewReason"] = (
            "Official recurring day/month dates are published separately with a "
            "GradWindow-mapped cycle year; they are not exact official-cycle records."
        )
    else:
        candidate["id"] = f"known-programme-guidance:{programme.id}"
        candidate["type"] = "known-programme-window-guidance"
    return candidate


def _candidate_record(
    adapter,
    programme,
    shared_opens_at: str | None,
    detected_at: str,
) -> dict:
    shared_opening_basis = adapter.application_opens_at_basis

    def opening_for(window) -> tuple[str | None, str]:
        opens_at = window.opens_at or shared_opens_at
        if not opens_at or opens_at > window.closes_at:
            return None, "missing"
        if window.opens_at:
            return opens_at, window.opens_at_basis or "official"
        return opens_at, shared_opening_basis

    windows = []
    scope_type = getattr(adapter, "known_programme_window_scope_type", "programme")
    scope_id = getattr(adapter, "known_programme_window_scope_id", None) or programme.id
    for window in programme.windows:
        opens_at, opens_at_basis = opening_for(window)
        window_record = {
            "intake": window.intake or adapter.intake,
            "round": window.round,
            "applicantCategories": window.applicant_categories,
            "opensAt": opens_at,
            "opensAtBasis": opens_at_basis,
            "closesAt": window.closes_at,
            "sourceUrl": window.source_url or programme.source_url,
        }
        if scope_type != "programme" or scope_id != programme.id:
            window_record.update({"scopeType": scope_type, "scopeId": scope_id})
        windows.append(window_record)
    has_unresolved_opening = any(window["opensAt"] is None for window in windows)
    has_inferred_opening = any(
        window.get("opensAtBasis", "").startswith("inferred") for window in windows
    )
    has_recurring_policy = any(
        window.get("opensAtBasis") == "official-recurring-policy" for window in windows
    )
    deadline_precedes_shared_opening = bool(
        shared_opens_at
        and any(window["closesAt"] < shared_opens_at for window in windows)
    )
    candidate = {
        "id": f"new-programme:{programme.id}",
        "type": "new-programme",
        "status": "pending",
        "universityId": adapter.university_id,
        "detectedAt": detected_at,
        "sourceUrl": programme.source_url,
        "programme": {
            "id": programme.id,
            "universityId": adapter.university_id,
            "name": programme.name,
            "degreeType": programme.degree_type,
            "faculty": " | ".join(
                value for value in (programme.faculty, programme.department) if value
            ),
            "applicationUrl": programme.application_url,
            "sourceUrl": programme.source_url,
        },
        "windows": windows,
        "parseStatus": programme.parse_status,
        "reviewReason": (
            "No application deadline was parsed."
            if not windows
            else (
                (
                    "An early deadline precedes the shared commencement date; "
                    "confirm the programme-specific opening date."
                )
                if deadline_precedes_shared_opening
                else (
                    "At least one opening date is not published as an exact date; "
                    "confirm it on the programme page."
                )
                if has_unresolved_opening
                else (
                    "Dates were materialized from an official recurring schedule; "
                    "the cycle year is derived and is not eligible for automatic "
                    "publication."
                )
                if has_recurring_policy
                else (
                    "Opening date uses a configured cycle default; review the "
                    "officially parsed deadline before promotion."
                )
                if has_inferred_opening
                else "Review the automatically discovered programme and application rounds."
            )
        ),
        "evidenceExcerpt": programme.deadline_text,
    }
    if programme.retrieval_method or programme.evidence_quality:
        candidate["discoveryEvidence"] = {
            "retrievalMethod": programme.retrieval_method,
            "evidenceQuality": programme.evidence_quality,
            "documentHash": programme.evidence_document_hash,
        }
    return candidate


def _hash(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def _source_hash(content: str) -> str:
    """Fingerprint deadline signals, excluding unrelated page-content churn."""
    signal = deadline_signal_text(content, max_lines=None)
    normalised = re.sub(r"\s+", " ", signal).strip().lower()
    return _hash(normalised)


def _effective_opening(adapter, catalog, window) -> str | None:
    opens_at = window.opens_at or catalog.application_opens_at
    if not opens_at or opens_at > window.closes_at:
        return None
    return opens_at


def _is_official_exact_window(adapter, catalog, window) -> bool:
    return _has_official_exact_opening(adapter, catalog, window)


def _is_official_recurring_policy_window(adapter, catalog, window) -> bool:
    if not _effective_opening(adapter, catalog, window):
        return False
    opening_basis = window.opens_at_basis or (
        "official" if window.opens_at else adapter.application_opens_at_basis
    )
    return opening_basis == "official-recurring-policy"


def _has_official_exact_opening(adapter, catalog, window) -> bool:
    if not _effective_opening(adapter, catalog, window):
        return False
    opening_basis = window.opens_at_basis or (
        "official" if window.opens_at else adapter.application_opens_at_basis
    )
    return opening_basis == "official"


def _window_status(
    exact_window_count: int,
    recurring_policy_window_count: int,
    observed_window_count: int,
    missing_opening_date_count: int,
    programmes_without_deadlines: int,
) -> str:
    if (
        exact_window_count
        and exact_window_count == observed_window_count
        and not programmes_without_deadlines
    ):
        return "exact"
    accounted_window_count = exact_window_count + recurring_policy_window_count
    if (
        recurring_policy_window_count
        and accounted_window_count == observed_window_count
    ):
        if programmes_without_deadlines:
            return (
                "mixed-policy-partial"
                if exact_window_count
                else "recurring-policy-partial"
            )
        return "mixed-policy" if exact_window_count else "recurring-policy"
    if exact_window_count:
        return "partial"
    if observed_window_count or missing_opening_date_count:
        return "needs-opening-date"
    return "monitoring"


def _limitation_reason(
    exact_window_count: int,
    recurring_policy_window_count: int,
    observed_window_count: int,
    missing_opening_date_count: int,
    programmes_without_deadlines: int,
) -> str | None:
    if (
        exact_window_count
        and exact_window_count == observed_window_count
        and not programmes_without_deadlines
    ):
        return None
    if recurring_policy_window_count:
        parts = [
            f"{recurring_policy_window_count} window(s) use an official recurring "
            "day/month policy; GradWindow maps the cycle year."
        ]
        if missing_opening_date_count:
            parts.append(
                f"{missing_opening_date_count} additional parsed deadline(s) lack "
                "an official exact opening date."
            )
        if programmes_without_deadlines:
            parts.append(
                f"{programmes_without_deadlines} programme(s) expose no parsed deadline."
            )
        return " ".join(parts)
    if missing_opening_date_count:
        return (
            f"{missing_opening_date_count} parsed deadline(s) lack an official "
            "exact opening date."
        )
    if observed_window_count:
        suffix = (
            f" {programmes_without_deadlines} programme(s) expose no parsed deadline."
            if programmes_without_deadlines
            else ""
        )
        return (
            "Observed dates are incomplete or not safe for automatic publication."
            f"{suffix}"
        )
    return (
        "The checked official sources currently expose no complete exact opening-"
        "and-closing window to this adapter."
    )
